"""Seed `order_file/IO_Order.xlsx` from the legacy `Order_IO.xlsx`.

The new workbook has one sheet per machine type, each holding three sections
(front / circuit / back) for both directions. Circuit rows are written once
with `#` standing in for the circuit number.

The legacy file spells every circuit out, so the circuit block is derived by
taking circuit 1's rows and checking that substituting `#` reproduces
circuit 2's rows; mismatches are reported so they can be fixed by hand.

Usage:
    python scripts/build_order_file.py [--force]
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parent.parent
LEGACY_ORDER = REPO_ROOT / "legacy_code_interface" / "order_file" / "Order_IO.xlsx"
TARGET = REPO_ROOT / "order_file" / "IO_Order.xlsx"

HEADERS = ["Section", "Direction", "Order", "Label", "IOType", "Description"]

# Legacy sheet pairs -> machine type.
SOURCES = {
    "Regular": {"Input": "Input", "Output": "Output"},
    "AHU": {"Input": "Input (2)", "Output": "Output (2)"},
}

DEFAULT_IO_TYPE = {"Input": "non connecter", "Output": "non connecterO"}

# Labels whose trailing number counts globally, not per circuit, so a single
# `#` cannot express them.
OVERRIDES = {
    "ReservedO1": "ReservedO#-1",
    "ReservedO2": "ReservedO#-2",
    "RESERVED2": "Reserved#",
}

MACHINE_ROW = "N"  # legacy marker for "not tied to a circuit"


def _cell(row, index) -> str:
    if index >= len(row):
        return ""
    value = row.iloc[index]
    if pd.isna(value):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _templatize(label: str, circuit: str) -> str:
    if label in OVERRIDES:
        return OVERRIDES[label]
    return label.replace(circuit, "#") if circuit and circuit in label else label


def parse_sheet(path: Path, sheet: str, direction: str) -> tuple[list[dict], list[str]]:
    """Split one legacy sheet into front / circuit / back entries."""
    frame = pd.read_excel(path, sheet_name=sheet)
    warnings: list[str] = []

    front: list[dict] = []
    back: list[dict] = []
    by_circuit: dict[str, list[dict]] = {}
    seen_circuit = False

    for _, row in frame.iterrows():
        label = _cell(row, 1)
        circuit = _cell(row, 2)
        io_type = _cell(row, 3) or DEFAULT_IO_TYPE[direction]
        if not label or label.lower() == "nan":
            continue

        entry = {"label": label, "io_type": io_type}
        if circuit == MACHINE_ROW or not circuit:
            (back if seen_circuit else front).append(entry)
        else:
            seen_circuit = True
            by_circuit.setdefault(circuit, []).append(entry)

    circuit_keys = sorted(by_circuit, key=lambda k: int(k) if k.isdigit() else k)
    if not circuit_keys:
        return _numbered(front, "front", direction) + _numbered(back, "back", direction), warnings

    first, *rest = circuit_keys
    reference = by_circuit[first]
    second = by_circuit[rest[0]] if rest else []

    circuit_entries: list[dict] = []
    for index, entry in enumerate(reference):
        template = _templatize(entry["label"], first)
        if second and index < len(second):
            expected = template.replace("#", rest[0])
            if expected != second[index]["label"] and entry["label"] not in OVERRIDES:
                warnings.append(
                    f"{sheet}: '{entry['label']}' -> '{template}' would give "
                    f"'{expected}' for circuit {rest[0]}, but the file has "
                    f"'{second[index]['label']}'"
                )
        circuit_entries.append({"label": template, "io_type": entry["io_type"]})

    rows = (
        _numbered(front, "front", direction)
        + _numbered(circuit_entries, "circuit", direction)
        + _numbered(back, "back", direction)
    )
    return rows, warnings


def _numbered(entries: list[dict], section: str, direction: str) -> list[dict]:
    return [
        {
            "section": section,
            "direction": direction,
            "order": index + 1,
            "label": entry["label"],
            "io_type": entry["io_type"],
            "description": "",
        }
        for index, entry in enumerate(entries)
    ]


def write_workbook(sheets: dict[str, list[dict]], target: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="667EEA")
    section_fills = {
        "front": PatternFill("solid", fgColor="EEF2FF"),
        "circuit": PatternFill("solid", fgColor="FFFFFF"),
        "back": PatternFill("solid", fgColor="F5F3FF"),
    }

    for machine_type, rows in sheets.items():
        sheet = workbook.create_sheet(machine_type)
        sheet.append(HEADERS)
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row in rows:
            sheet.append([
                row["section"],
                row["direction"],
                row["order"],
                row["label"],
                row["io_type"],
                row["description"],
            ])
            for cell in sheet[sheet.max_row]:
                cell.fill = section_fills[row["section"]]

        for index, width in enumerate((10, 11, 8, 22, 18, 26), start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:F{sheet.max_row}"

    target.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(target)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--force", action="store_true", help="overwrite an existing file")
    args = parser.parse_args()

    if TARGET.exists() and not args.force:
        print(f"[SKIP] {TARGET} already exists (use --force to regenerate).")
        return 0
    if not LEGACY_ORDER.exists():
        print(f"[ERROR] Legacy order file not found: {LEGACY_ORDER}", file=sys.stderr)
        return 1

    sheets: dict[str, list[dict]] = {}
    all_warnings: list[str] = []
    for machine_type, mapping in SOURCES.items():
        rows: list[dict] = []
        for direction, sheet_name in mapping.items():
            parsed, warnings = parse_sheet(LEGACY_ORDER, sheet_name, direction)
            rows.extend(parsed)
            all_warnings.extend(warnings)
        sheets[machine_type] = rows

    write_workbook(sheets, TARGET)

    print(f"[OK] Wrote {TARGET}")
    for machine_type, rows in sheets.items():
        counts = {}
        for row in rows:
            counts[(row["section"], row["direction"])] = counts.get((row["section"], row["direction"]), 0) + 1
        summary = ", ".join(f"{s}/{d}={n}" for (s, d), n in sorted(counts.items()))
        print(f"  {machine_type}: {summary}")

    for warning in all_warnings:
        print(f"  [WARN] {warning}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
