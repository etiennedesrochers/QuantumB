"""
Workbook helpers for the XNNOV-RS-Database.xlsm data source.

The web UI uses the Tables sheet as the source of truth for compressor
selection data. This module exposes a small read-only API that returns the
sheet contents in a JSON-friendly shape.
"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
DEFAULT_WORKBOOK_PATH = PROJECT_ROOT / "data_excel" / "XNNOV-RS-Database.xlsm"


def _clean_value(value: Any) -> Any:
    if isinstance(value, str):
        return value.strip()
    return value


def _is_empty_row(row: tuple[Any, ...]) -> bool:
    return all(value is None or str(value).strip() == "" for value in row)


def _find_header_row(rows: list[tuple[Any, ...]]) -> int:
    expected_markers = {
        "Nominal Capacity",
        "Circuit_ID",
        "Compressor_Model",
        "Compressor Model 200V",
    }

    for index, row in enumerate(rows):
        normalized = {str(cell).strip() for cell in row if cell is not None}
        if normalized.intersection(expected_markers):
            return index

    raise ValueError("Could not find a header row on the Tables sheet")


def _worksheet_rows_for_ref(worksheet, ref: str) -> list[tuple[Any, ...]]:
    return [tuple(cell.value for cell in row) for row in worksheet[ref]]


def _table_to_dict(worksheet, table_name: str) -> dict[str, Any]:
    table = worksheet.tables[table_name]
    ref = getattr(table, "ref", table)
    raw_rows = _worksheet_rows_for_ref(worksheet, ref)
    if not raw_rows:
        return {
            "name": table_name,
            "ref": ref,
            "headers": [],
            "rows": [],
        }

    headers = [
        str(cell).strip()
        for cell in raw_rows[0]
        if cell is not None and str(cell).strip() != ""
    ]

    rows: list[dict[str, Any]] = []
    for raw_row in raw_rows[1:]:
        if _is_empty_row(raw_row):
            continue

        record: dict[str, Any] = {}
        for column_index, header in enumerate(headers):
            record[header] = _clean_value(raw_row[column_index]) if column_index < len(raw_row) else None
        rows.append(record)

    return {
        "name": table_name,
        "ref": ref,
        "headers": headers,
        "rows": rows,
    }


def load_table_catalog(workbook_path: Path | None = None) -> dict[str, Any]:
    """Return the named Excel tables that live on the Tables worksheet."""
    path = workbook_path or DEFAULT_WORKBOOK_PATH
    if not path.exists():
        return {
            "workbook": str(path),
            "sheet": "Tables",
            "tables": [],
            "error": "Workbook not found",
        }

    workbook = load_workbook(path, data_only=True, keep_vba=True)
    if "Tables" not in workbook.sheetnames:
        return {
            "workbook": str(path),
            "sheet": "Tables",
            "tables": [],
            "error": "Tables sheet not found",
        }

    worksheet = workbook["Tables"]
    tables = [_table_to_dict(worksheet, name) for name in worksheet.tables.keys()]

    return {
        "workbook": str(path),
        "sheet": "Tables",
        "tables": tables,
    }


def load_table_rows(table_name: str, workbook_path: Path | None = None) -> list[dict[str, Any]]:
    """Return rows for a specific named table on the Tables worksheet."""
    path = workbook_path or DEFAULT_WORKBOOK_PATH
    workbook = load_workbook(path, data_only=True, keep_vba=True)
    worksheet = workbook["Tables"]
    return _table_to_dict(worksheet, table_name)["rows"]


def load_compressor_rows(workbook_path: Path | None = None) -> list[dict[str, Any]]:
    """Return compressor selection rows from the SkidDB table."""
    rows = load_table_rows("SkidDB", workbook_path)

    compressors: list[dict[str, Any]] = []
    for row in rows:
        compressors.append(
            {
                "nominal_capacity": row.get("Nominal Capacity"),
                "manufacturer": row.get("Compressor Manufacturer"),
                "control": row.get("Control"),
                "refrigerant": row.get("Refrigerant"),
                "skid_model_number": row.get("Compressor Skid Model Number"),
                "compressor_1_qty": row.get("Compressor 1 Qty"),
                "compressor_2_qty": row.get("Compressor 2 Qty"),
                "minimum_capacity": row.get("Minimum Capacity"),
                "models_by_voltage": {
                    "200v": row.get("Compressor Model 200V"),
                    "400v": row.get("Compressor Model 400V"),
                    "600v": row.get("Compressor Model 600V"),
                },
                "compressor_qty": row.get("Compressor Qty"),
                "condenser_airflow": row.get("Condenser Airflow"),
                "suction_line_dia": row.get("Suction Line Dia"),
                "discharge_line_dia": row.get("Discharge Line Dia"),
                "liquid_line_dia": row.get(" Liquide Line Dia"),
                "cooling_air_condenser_model": row.get("Cooling Air Condenser Model"),
                "hp_air_condenser_model": row.get("HP  Air Condenser Model"),
                "cooling_water_condenser_model": row.get("Cooling Water Condenser Model"),
                "hp_water_condenser_model": row.get("HP  Ater Condenser Model"),
                "water_evaporator": row.get("Water Evaporator"),
                "condenser_fan_qty": row.get("Condenser Fan Qty"),
                "raw": row,
            }
        )

    return compressors


def _to_number(value: Any) -> float | None:
    try:
        if value is None or value == "":
            return None
        return float(value)
    except Exception:
        return None


def _to_int(value: Any, default: int = 0) -> int:
    try:
        if value is None or value == "":
            return default
        return int(float(value))
    except Exception:
        return default


def _voltage_to_key(tension: str) -> str:
    normalized = str(tension).strip().lower().replace(" ", "")
    mapping = {
        "200": "200v",
        "208": "200v",
        "200v": "200v",
        "400": "400v",
        "480": "400v",
        "400v": "400v",
        "600": "600v",
        "575": "600v",
        "600v": "600v",
    }
    return mapping.get(normalized, "400v")


def load_generator_filters(
    workbook_path: Path | None = None,
    manufacturer: str | None = None,
) -> dict[str, Any]:
    """Return valid manufacturer and nominal capacity (tons) choices for generator flow.

    Source of truth: CondUnitDB table (starts on row 98 in the workbook).
    """
    cond_rows = load_table_rows("CondUnitDB", workbook_path)

    manufacturers = sorted(
        {
            str(row.get("Compressor Manufacturer", "")).strip()
            for row in cond_rows
            if str(row.get("Compressor Manufacturer", "")).strip()
        }
    )

    selected_manufacturer = (manufacturer or "").strip().lower()
    capacities = sorted(
        {
            _to_number(row.get("Nominal Capacity"))
            for row in cond_rows
            if _to_number(row.get("Nominal Capacity")) is not None
            and (
                not selected_manufacturer
                or str(row.get("Compressor Manufacturer", "")).strip().lower() == selected_manufacturer
            )
        }
    )

    return {
        "capacities": capacities,
        "manufacturers": manufacturers,
        "capacity_unit": "tons",
        "tensions": [
            {"label": "208 V", "value": "208"},
            {"label": "480 V", "value": "480"},
            {"label": "575 V", "value": "575"},
        ],
    }


def load_circuits_for_selection(
    capacity: float,
    manufacturer: str,
    tension: str,
    workbook_path: Path | None = None,
) -> dict[str, Any]:
    """Return circuits and compressors for a capacity/manufacturer/tension selection."""
    voltage_key = _voltage_to_key(tension)
    skid_rows = load_compressor_rows(workbook_path)
    cond_rows = load_table_rows("CondUnitDB", workbook_path)

    manufacturer_norm = str(manufacturer).strip().lower()
    capacity_num = _to_number(capacity)

    skid_lookup: dict[str, dict[str, Any]] = {}
    for row in skid_rows:
        skid_id = str(row.get("skid_model_number") or "").strip()
        if skid_id:
            skid_lookup[skid_id] = row

    matching_cond = []
    for row in cond_rows:
        row_capacity = _to_number(row.get("Nominal Capacity"))
        row_manufacturer = str(row.get("Compressor Manufacturer", "")).strip().lower()
        if row_capacity == capacity_num and row_manufacturer == manufacturer_norm:
            matching_cond.append(row)

    circuits: list[dict[str, Any]] = []

    if matching_cond:
        for row in matching_cond:
            circuit_qty = max(1, _to_int(row.get("Circuit Qty"), default=1))
            compressor_skids: list[str] = []
            circuit_specific_skids: dict[int, list[str]] = {}
            has_circuit_prefixed_columns = False

            for column_name in row.keys():
                if "Compressor Skid" not in column_name:
                    continue
                value = row.get(column_name)
                skid_id = str(value).strip() if value is not None else ""
                if not skid_id or skid_id == "0":
                    continue
                compressor_skids.append(skid_id)

                # CondUnitDB uses headers like "C1 Compressor Skid" / "C2 Compressor Skid2".
                # When present, map each skid only to its target circuit index.
                match = re.search(r"^\s*C\s*(\d+)\b", str(column_name), flags=re.IGNORECASE)
                if match:
                    has_circuit_prefixed_columns = True
                    circuit_index = max(1, _to_int(match.group(1), default=1))
                    circuit_specific_skids.setdefault(circuit_index, []).append(skid_id)

            if not compressor_skids:
                continue

            for circuit_index in range(1, circuit_qty + 1):
                compressors: list[dict[str, Any]] = []
                skids_for_circuit = (
                    circuit_specific_skids.get(circuit_index, [])
                    if has_circuit_prefixed_columns
                    else compressor_skids
                )

                for skid_id in skids_for_circuit:
                    skid = skid_lookup.get(skid_id)
                    if not skid:
                        continue
                    model = (
                        (skid.get("models_by_voltage") or {}).get(voltage_key)
                        or skid.get("skid_model_number")
                        or skid_id
                    )
                    compressors.append(
                        {
                            "model_number": model,
                            "description": skid_id,
                            "skid_model_number": skid_id,
                            "quantity": max(1, _to_int(skid.get("compressor_qty"), default=1)),
                            "templates": [],
                        }
                    )

                if compressors:
                    circuits.append(
                        {
                            "name": f"CU{len(circuits)+1:03d}",
                            "description": f"{manufacturer} {capacity_num:g} tons Circuit {len(circuits)+1}",
                            "compressors": compressors,
                        }
                    )

    if not circuits:
        fallback_rows = [
            row
            for row in skid_rows
            if _to_number(row.get("nominal_capacity")) == capacity_num
            and str(row.get("manufacturer", "")).strip().lower() == manufacturer_norm
        ]
        if fallback_rows:
            compressors = []
            for row in fallback_rows:
                model = (
                    (row.get("models_by_voltage") or {}).get(voltage_key)
                    or row.get("skid_model_number")
                )
                compressors.append(
                    {
                        "model_number": model,
                        "description": row.get("skid_model_number"),
                        "skid_model_number": row.get("skid_model_number"),
                        "quantity": max(1, _to_int(row.get("compressor_qty"), default=1)),
                        "templates": [],
                    }
                )

            circuits.append(
                {
                    "name": "CU001",
                    "description": f"{manufacturer} {capacity_num:g} tons",
                    "compressors": compressors,
                }
            )

    return {
        "capacity": capacity_num,
        "manufacturer": manufacturer,
        "tension": str(tension),
        "voltage_key": voltage_key,
        "circuits": circuits,
    }