"""The I/O ordering workbook (`order_file/IO_Order.xlsx`).

One sheet per machine type, each holding three sections in this order:

    front    machine-level I/O that is always present, before the circuits
    circuit  the block repeated once per circuit; `#` = the circuit number
    back     machine-level I/O that closes the list

Columns: Section | Direction | Order | Label | IOType | Description.

Entries the templates do not produce are injected as "Reserved" placeholders,
matching what the desktop GUI does with the legacy `Order_IO.xlsx`.
"""

from __future__ import annotations

import dataclasses
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..legacy_bridge import REPO_ROOT, ensure_legacy_importable
from .errors import NotFoundError, ServiceError

ensure_legacy_importable()

from src.core.io_manager import IOItem  # noqa: E402

ORDER_PATH = REPO_ROOT / "order_file" / "IO_Order.xlsx"

SECTIONS = ("front", "circuit", "back")
DIRECTIONS = ("Input", "Output")
CIRCUIT_PLACEHOLDER = "#"
RESERVED_DESCRIPTION = "Reserved"

_DEFAULT_IO_TYPE = {"Input": "non connecter", "Output": "non connecterO"}
_COLUMNS = ("section", "direction", "order", "label", "io_type", "description")


def _text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _open_workbook(**kwargs):
    if not ORDER_PATH.exists():
        raise ServiceError(f"Ordering workbook not found: {ORDER_PATH}")
    try:
        return load_workbook(ORDER_PATH, **kwargs)
    except PermissionError as exc:
        raise ServiceError(
            f"Cannot read {ORDER_PATH.name} — it is locked by another program "
            "(close it in Excel and retry)."
        ) from exc
    except OSError as exc:
        raise ServiceError(f"Cannot read {ORDER_PATH.name}: {exc}") from exc


def _read_sheet(machine_type: str) -> list[dict[str, Any]]:
    workbook = _open_workbook(read_only=True, data_only=True)
    try:
        if machine_type not in workbook.sheetnames:
            raise NotFoundError(f"No ordering sheet for machine type '{machine_type}'")
        sheet = workbook[machine_type]
        rows: list[dict[str, Any]] = []
        for index, raw in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
            row = dict(zip(_COLUMNS, (_text(value) for value in raw)))
            if not row.get("label"):
                continue

            section = row["section"].lower()
            direction = row["direction"].capitalize()
            if section not in SECTIONS or direction not in DIRECTIONS:
                raise ServiceError(
                    f"{machine_type} row {index + 2}: invalid section/direction "
                    f"'{row['section']}/{row['direction']}'"
                )

            rows.append(
                {
                    "section": section,
                    "direction": direction,
                    "order": int(row["order"]) if row["order"].isdigit() else index,
                    "label": row["label"],
                    "io_type": row["io_type"] or _DEFAULT_IO_TYPE[direction],
                    "description": row["description"],
                }
            )
        return rows
    finally:
        workbook.close()


def list_machine_types() -> list[str]:
    """Machine types are the sheets of the ordering workbook."""
    if not ORDER_PATH.exists():
        return []
    workbook = _open_workbook(read_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def load_order(machine_type: str) -> dict[str, Any]:
    """Return one machine type's rows, grouped by section and direction."""
    rows = _read_sheet(machine_type)
    grouped = {
        section: {
            direction: sorted(
                (r for r in rows if r["section"] == section and r["direction"] == direction),
                key=lambda r: r["order"],
            )
            for direction in DIRECTIONS
        }
        for section in SECTIONS
    }
    return {
        "machine_type": machine_type,
        "sections": grouped,
        "counts": {
            direction: sum(len(grouped[s][direction]) for s in SECTIONS)
            for direction in DIRECTIONS
        },
    }


def expand(machine_type: str, circuit_numbers: list[str]) -> list[dict[str, Any]]:
    """Flatten the sheet into the expected I/O sequence for these circuits.

    Inputs come first, then outputs; within each direction the order is
    front, then the circuit block once per circuit, then back.
    """
    grouped = load_order(machine_type)["sections"]
    expanded: list[dict[str, Any]] = []

    for direction in DIRECTIONS:
        for entry in grouped["front"][direction]:
            expanded.append(_entry(entry, circuit_no=""))
        for circuit_no in circuit_numbers:
            for entry in grouped["circuit"][direction]:
                expanded.append(_entry(entry, circuit_no=circuit_no))
        for entry in grouped["back"][direction]:
            expanded.append(_entry(entry, circuit_no=""))

    return expanded


def _entry(entry: dict[str, Any], circuit_no: str) -> dict[str, Any]:
    return {
        "section": entry["section"],
        "direction": entry["direction"],
        "tag": entry["label"].replace(CIRCUIT_PLACEHOLDER, circuit_no),
        "io_type": entry["io_type"],
        "description": entry["description"],
        "circuit_no": circuit_no,
    }


def apply_order(
    items: list[IOItem],
    machine_type: str,
    circuit_numbers: list[str],
    circuit_names: dict[str, str] | None = None,
) -> list[IOItem]:
    """Sort *items* into the workbook order, injecting missing entries.

    Items the workbook does not mention keep their relative order and are
    appended after the ordered ones, per direction.
    """
    expected = expand(machine_type, circuit_numbers)
    names = circuit_names or {}

    remaining = {}
    for item in items:
        remaining.setdefault(item.tag.strip().lower(), []).append(item)

    ordered: list[IOItem] = []
    for entry in expected:
        matches = remaining.get(entry["tag"].lower())
        if matches:
            ordered.append(matches.pop(0))
        else:
            ordered.append(
                IOItem(
                    tag=entry["tag"],
                    io_type=entry["direction"],
                    description=entry["description"] or RESERVED_DESCRIPTION,
                    signal_type="Analog",
                    io_type_name=entry["io_type"],
                    old_name=entry["tag"],
                    old_description=entry["description"] or RESERVED_DESCRIPTION,
                    circuit_no=entry["circuit_no"],
                    circuit_name=names.get(entry["circuit_no"], ""),
                )
            )

    leftovers = [item for group in remaining.values() for item in group]
    inputs = [i for i in leftovers if i.io_type == "Input"]
    outputs = [i for i in leftovers if i.io_type != "Input"]

    return (
        [i for i in ordered if i.io_type == "Input"]
        + inputs
        + [i for i in ordered if i.io_type != "Input"]
        + outputs
    )


def clone_items(items: list[IOItem]) -> list[IOItem]:
    """Copies, so ordering never mutates the caller's list."""
    return [dataclasses.replace(item) for item in items]
